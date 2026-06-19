"""
LeadForge AI — Export Service
Generates CSV and Excel exports of leads with all related data.
"""
import os
import csv
import uuid
from typing import List, Optional, Dict, Any
from datetime import datetime

from app.config import settings
from app.infrastructure.database.models import Lead


def _flatten_lead(lead: Lead) -> Dict[str, Any]:
    """Flatten a lead and all its relations into a single dict for export."""
    # Primary contact
    primary_contact = lead.contact_info[0] if lead.contact_info else None
    # All emails
    emails = [c.email for c in lead.contact_info if c.email]
    phones = [c.phone for c in lead.contact_info if c.phone]

    # Executive
    primary_exec = lead.executive_info[0] if lead.executive_info else None

    # Website analysis
    wa = lead.website_analysis

    # Tags
    tags = ", ".join([t.name for t in lead.tags])

    return {
        "Company Name": lead.company_name,
        "Industry": lead.industry,
        "Country": lead.country,
        "State": lead.state,
        "City": lead.city,
        "Address": lead.address,
        "Website": lead.website,
        "Email(s)": " | ".join(emails),
        "Phone(s)": " | ".join(phones),
        "Facebook": primary_contact.facebook_url if primary_contact else "",
        "Instagram": primary_contact.instagram_url if primary_contact else "",
        "LinkedIn": primary_contact.linkedin_url if primary_contact else "",
        "Twitter/X": primary_contact.twitter_url if primary_contact else "",
        "YouTube": primary_contact.youtube_url if primary_contact else "",
        "Executive Name": primary_exec.executive_name if primary_exec else "",
        "Executive Position": primary_exec.position if primary_exec else "",
        "Executive LinkedIn": primary_exec.linkedin_profile if primary_exec else "",
        "Rating": lead.rating,
        "Review Count": lead.review_count,
        "Lead Score": lead.lead_score,
        "Website Score": lead.website_score,
        "Status": lead.status.value if lead.status else "",
        "Tags": tags,
        "SSL Enabled": wa.ssl_enabled if wa else "",
        "Mobile Responsive": wa.mobile_responsive if wa else "",
        "Has Contact Form": wa.has_contact_form if wa else "",
        "Has Social Links": wa.has_social_links if wa else "",
        "Improvement Summary": wa.improvement_summary if wa else "",
        "Opportunity Notes": lead.opportunity_insight.ai_notes if lead.opportunity_insight else "",
        "Source": lead.source,
        "Date Added": lead.created_at.strftime("%Y-%m-%d") if lead.created_at else "",
    }


async def export_to_csv(leads: List[Lead], filename: Optional[str] = None) -> str:
    """Export leads to CSV file. Returns the file path."""
    os.makedirs(settings.export_dir, exist_ok=True)
    fname = filename or f"leadforge_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}.csv"
    filepath = os.path.join(settings.export_dir, fname)

    rows = [_flatten_lead(lead) for lead in leads]
    if not rows:
        return filepath

    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)

    return filepath


async def export_to_xlsx(leads: List[Lead], filename: Optional[str] = None) -> str:
    """Export leads to Excel XLSX file. Returns the file path."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    os.makedirs(settings.export_dir, exist_ok=True)
    fname = filename or f"leadforge_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}.xlsx"
    filepath = os.path.join(settings.export_dir, fname)

    rows = [_flatten_lead(lead) for lead in leads]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LeadForge AI Leads"

    if not rows:
        wb.save(filepath)
        return filepath

    headers = list(rows[0].keys())

    # Style the header row
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Write data rows
    row_fill_1 = PatternFill(start_color="F8F9FF", end_color="F8F9FF", fill_type="solid")
    row_fill_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for row_idx, row_data in enumerate(rows, 2):
        fill = row_fill_1 if row_idx % 2 == 0 else row_fill_2
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    # Auto-size columns (rough estimate)
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 40)

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    wb.save(filepath)
    return filepath
